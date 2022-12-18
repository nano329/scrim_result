import meta
from copy import deepcopy
from collections import Counter
import os
import KartRider as KR
from openpyxl import Workbook
import xlsxwriter
from openpyxl.styles import Color, PatternFill, Font, Alignment


setPath = False
if os.path.isdir("./metadata"):
    print("asdf")
    setPath = True
    KR.set_metadatapath("./metadata")
api = KR.Api(meta.api_key)
arr = [(0, 0), (0, 1), (0, 2), (0, 3), (1, 0), (1, 1), (1, 2), (1, 3)]
f = Font(color="ff0000")
RED = f
YELLOW = Font(color="ffff00")
BLUE = Font(color="0000ff")


def MemberNick(id):
    user = KR.User(api, accessid=meta.member_accessId[id])
    return user.name


def avg(lst, ret):
    s = 0
    for i in lst:
        if i == -1: s += ret
        else: s += i
    return (s * 100) / len(lst)


def SetTime(s):
    day = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    s = s.replace('-', ' ').replace(':', ' ')
    arr = list(map(int, s.split(' ')))
    arr[3] -= 9
    if arr[3] < 0:
        arr[3] += 24
        arr[2] -= 1
        if arr[2] <= 0:
            if arr[1] != 1:
                arr[2] += day[arr[1] - 2]
                if arr[0] % 400 == 0 or (arr[0] % 100 != 0 and arr[0] % 4 == 0):
                    if arr[1] == 3:
                        arr[2] += 1
                arr[1] -= 1
            else:
                arr[0] -= 1
                arr[1] = 12
                arr[2] = 31

    for i in range(1, 6):
        arr[i] = str(arr[i])
        if len(arr[i]) == 1:
            arr[i] = '0' + arr[i]

    return str(arr[0]) + '-' + arr[1] + '-' + arr[2] + ' ' + arr[3] + ':' + arr[4] + ':' + arr[5]


def StrToDate(s):
    day = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day2 = [31,29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    s = s.replace('-', ' ').replace(':', ' ')
    arr = list(map(int, s.split(' ')))
    date = 0
    if arr[0] % 4 == 0:
        date = (sum(day2[0:arr[1]-1]) + arr[2]) * 24 * 60 + arr[3] * 60 + arr[4]
    else:
        date = (sum(day[0:arr[1] -1]) + arr[2]) * 24 * 60 + arr[3] * 60 + arr[4]
    return date


def DateToStr(date):
    return


def IntToRow(n):
    if n <= 26:
        return chr(n + 64)
    else:
        return chr(n // 26 + 64) + chr(n % 26 + 64)


def record(n):
    sec = n // 1000
    msec = n % 1000
    mi = sec // 60
    sec -= mi * 60
    mi = str(mi)
    sec = '{0:02d}'.format(sec)
    msec = '{0:03d}'.format(msec)
    if n >= 60000:
        return mi + "'" + sec + "'" + msec
    else:
        if sec[0] == '0':
            sec = sec[1:]
        return sec + "'" + msec

def team_search(nick, start, end, team1, team2, ret=9, dir="."):
    start = SetTime(start)
    end = SetTime(end)
    print(".")
    speed_match = api.user(nick).getMatches(limit=50, match_types="스피드 팀전", start_date=start, end_date=end)
    print(",")
    item_match_get = api.user(nick).getMatches(limit=50, start_date=start, end_date=end)
    print(item_match_get)
    item_match = {}
    if 'Unknown' in item_match_get.keys(): item_match['Unknown'] = item_match_get['Unknown']
    if '아이템 팀전' in item_match_get.keys(): item_match['아이템 팀전'] = item_match_get['아이템 팀전']
    print(".")



    print(speed_match)
    print(item_match)
    speed_rank = [[[], [], [], []], [[], [], [], []]]
    speed_kart = [[[], [], [], []], [[], [], [], []]]
    speed_main = [[], []]
    speed_track = []
    speed_win = []
    speed_record = [[[], [], [], []], [[], [], [], []]]
    item_rank = [[[], [], [], []], [[], [], [], []]]
    item_kart = [[[], [], [], []], [[], [], [], []]]
    item_main = [[], []]
    item_track = []
    item_win = []
    item_record = [[[], [], [], []], [[], [], [], []]]
    for gametype, match in speed_match.items():
        for game in match:
            print(len(game.detail.teams[0]))
            if len(game.detail.teams[0]) != 4:
                continue
            for i, j in arr:
                player = game.detail.teams[i][j]
                if len(speed_rank[i][j]) == 0:
                    speed_rank[i][j].insert(1, player.charactername)
                speed_rank[i][j].insert(1, player.matchrank)
                speed_kart[i][j].insert(0, player.kart)
                speed_record[i][j].insert(0, player.matchtime)
                print(".", end='')
            speed_track.insert(0, game.detail.track)
            speed_win.insert(0, game.detail.matchresult)
    print("asdfasdfasdf")

    for gametype, match in item_match.items():
        for game in match:
            print(len(game.detail.teams[0]))
            if len(game.detail.teams[0]) != 4:
                continue
            for i, j in arr:
                player = game.detail.teams[i][j]
                if len(item_rank[i][j]) == 0:
                    item_rank[i][j].insert(1, player.charactername)
                item_rank[i][j].insert(1, player.matchrank)
                item_kart[i][j].insert(0, player.kart)
                item_record[i][j].insert(0, player.matchtime)

            item_track.insert(0, game.detail.track)
            item_win.insert(0, game.detail.matchresult)

    speed_1st = []
    item_1st = []
    print(speed_rank)
    print(item_rank)
    print(speed_win)
    print(speed_kart)
    print(item_kart)
    print(speed_record)
    print(item_record)

    for i in range(len(speed_record[0][0])):
        lst = []
        for j in range(2):
            for k in range(4):
                if speed_record[j][k][i] != 0:
                    lst.append(speed_record[j][k][i])
        speed_1st.append(min(lst))
    for i in range(len(item_record[0][0])):
        lst = []
        for j in range(2):
            for k in range(4):
                if item_record[j][k][i] != 0:
                    lst.append(item_record[j][k][i])
        item_1st.append(min(lst))
    print(item_1st)

    for i in range(2):
        for j in range(4):
            if len(speed_track) > 0:
                speed_main[i].append((Counter(speed_kart[i][j]).most_common(1)[0][0]))
            if len(item_track) > 0:
                item_main[i].append((Counter(item_kart[i][j]).most_common(1)[0][0]))

    print(speed_main)
    print(item_main)

    for i in range(4):
        if speed_rank[1][i][0] == nick:
            team1, team2 = team2, team1




    wb = xlsxwriter.Workbook(f'{dir}/{start.split()[0].replace("-", ".")} {team1} vs {team2}.xlsx')
    print(f'{dir}/{start.split()[0].replace("-", ".")} {team1} vs {team2}.xlsx')
    #Fix later

    dark_format = wb.add_format({"bg_color": "#000000", "font_color": "#ffffff", "font_size": 9, "align":"center", "valign":"center"})
    title = wb.add_format({"bg_color": "#000000", "font_color": "#ffffff", "font_size": 24, "align":"center", "valign":"center"})
    red_team = wb.add_format({"bg_color": "#c00000", "font_color": "#ffffff", "font_size": 16, "align":"center", "valign":"center"})
    blue_team = wb.add_format({"bg_color": "#002060", "font_color": "#ffffff", "font_size": 16, "align":"center", "valign":"center"})
    red_win = wb.add_format({"bg_color": "#f4b084", "font_color": "#000000", "font_size": 12, "align":"center", "valign":"center"})
    red_1st = wb.add_format({"bg_color": "#ed7d31", "font_color": "#ffffff", "font_size": 12, "bold": True, "align":"center", "valign":"center"})
    red_win_r = wb.add_format({"bg_color": "#f4b084", "font_color": "#ff0000", "font_size": 12, "bold": True, "align":"center", "valign":"center"})
    red_track = wb.add_format({"bg_color": "#ed7d31", "font_color": "#000000", "font_size": 12, "bold": True, "align":"center", "valign":"center"})
    blue_win = wb.add_format({"bg_color": "#8ea9db", "font_color": "#000000", "font_size": 12, "align":"center", "valign":"center"})
    blue_win_r = wb.add_format({"bg_color": "#8ea9db", "font_color": "#ff0000", "font_size": 12, "bold": True, "align":"center", "valign":"center"})
    blue_1st = wb.add_format({"bg_color": "#4472c4", "font_color": "#ffffff", "font_size": 12, "bold": True, "align":"center", "valign":"center"})
    blue_track = wb.add_format({"bg_color": "#4472c4", "font_color": "#000000", "font_size": 12, "bold": True, "align":"center", "valign":"center"})
    lose = wb.add_format({"bg_color": "#808080", "font_color": "#000000", "font_size": 12, "align":"center", "valign":"center"})
    lose_r = wb.add_format({"bg_color": "#808080", "font_color": "#f00000", "font_size": 12, "bold": True, "align":"center", "valign":"center"})

    if len(speed_track):
        ws = wb.add_worksheet(f'{start.split()[0].replace("-", ".")} {team1} vs {team2} SPEED')
        ws.set_column('A:H', 2)
        ws.set_column('I:I', 45)
        ws.set_column('J:Q', 2)
        for i in "BDFHKMOQ":
            ws.set_column(i + ':' + i, 12)

        ws.merge_range('A1:H1', team1, blue_team)
        ws.write('I1', start.split()[0], title)
        ws.write('I2', 'SPEED', dark_format)
        ws.merge_range('J1:Q1', team2, red_team)
        for i in [2, len(speed_rank[0][0])+2, len(speed_rank[0][0])+3]:
            for j in [65, 67, 69, 71, 74, 76, 78, 80]:
                ws.merge_range(chr(j) + str(i) + ':' + chr(j+1) + str(i), '', dark_format)
        for i, j in arr:
            player_list = speed_rank[i][j]
            ws.write(1, i * 9 + j * 2, player_list[0], dark_format)
            for m in range(1, len(player_list)):
                apply_format = None
                if str(i+1) != speed_win[m-1]: # win
                    if i == 0:
                        apply_format = blue_1st if player_list[m] == 1 else blue_win_r if player_list[m] == -1 else blue_win
                    else:
                        apply_format = red_1st if player_list[m] == 1 else red_win_r if player_list[m] == -1 else red_win
                else:
                    apply_format = lose_r if player_list[m] == -1 else lose
                ws.write(m + 1, i * 9 + j * 2, 9 if player_list[m] == -1 else player_list[m], apply_format)
                apply_format.set_font_size(8)
                ws.write(m + 1, i * 9 + j * 2 + 1, speed_kart[i][j][m-1], apply_format)

            ws.write(len(player_list) + 1, i * 9 + j * 2, speed_main[i][j], dark_format)
            ws.write(len(player_list) + 2, i * 9 + j * 2, str(int(avg(player_list[1:], ret)) / 100) + '등', dark_format)
        ws.write(len(speed_rank[0][0]) + 1, 8, "메인 카트", dark_format)
        ws.write(len(speed_rank[0][0]) + 2, 8, "평균 순위", dark_format)
        for i in range(len(speed_track)):
            apply_format = red_track if speed_win[i] == '1' else blue_track
            ws.write(i + 2, 8, speed_track[i] + f'({record(speed_1st[i])})', apply_format)


    if len(item_track):
        ws = wb.add_worksheet(f'{start.split()[0].replace("-", ".")} {team1} vs {team2} ITEM')
        ws.set_column('A:L', 2)
        ws.set_column('M:M', 45)
        ws.set_column('N:Y', 2)
        for i in "BEHKORUX":
            ws.set_column(i + ':' + i, 12)
        for i in "CFILPSVY":
            ws.set_column(i + ':' + i, 6)

        ws.merge_range('A1:L1', team1, blue_team)
        ws.write('M1', start.split()[0], title)
        ws.write('M2', 'ITEM', dark_format)
        ws.merge_range('N1:Y1', team2, red_team)
        for i in [2, len(item_rank[0][0]) + 2]:
            for j in [65, 68, 71, 74, 78, 81, 84, 87]:
                ws.merge_range(chr(j) + str(i) + ':' + chr(j + 2) + str(i), '', dark_format)
        for i, j in arr:
            player_list = item_rank[i][j]
            player_kart = item_kart[i][j]
            print(player_list)
            print(player_kart)
            print(len(player_list), len(item_1st))
            ws.write(1, i * 13 + j * 3, player_list[0], dark_format)
            #ws["abcdfghi"[i * 4 + j] + "1"] = player_list[0]
            for m in range(1, len(player_list)):
                apply_format = None
                if str(i + 1) != item_win[m - 1]:  # win
                    if i == 0:
                        apply_format = blue_1st if player_list[m] == 1 else blue_win
                    else:
                        apply_format = red_1st if player_list[m] == 1 else red_win
                else:
                    apply_format = lose
                apply_format.set_font_size(8)
                ws.write(m + 1, i * 13 + j * 3, 9 if player_list[m] == -1 else player_list[m], apply_format)
                ws.write(m + 1, i * 13 + j * 3 + 1, player_kart[m-1], apply_format)
                ws.write(m + 1, i * 13 + j * 3 + 2, f'({"Retire" if item_record[i][j][m-1] == 0 else "+" + str(record(item_record[i][j][m-1] - item_1st[m-1]))})', apply_format)
            ws.write(len(player_list) + 1, i * 12 + j * 3, item_main[i][j], dark_format)
        ws.write(len(item_rank[0][0]) + 1, 12, "메인 카트", dark_format)

        for i in range(len(item_track)):
            ws.write(i + 2, 12, item_track[i] + f'({record(item_1st[i])})', red_track if item_win[i] == '1' else blue_track)

    wb.close()


def point(arr):
    s = 0
    for i in arr:
        s += [-5, 10, 7, 5, 4, 3, 1, 0, -1, -5][i]
    return s


def item_point(arr):
    s = 0
    for i in arr:
        s += [-5, 10, 7, 5, 4, 3, 1, 0, -1, -5][i]
    return s


def indv_search(nick, start, end):
    start = SetTime(start)
    end = SetTime(end)
    print("Asdf")
    match_list = api.user(nick).getMatches(limit=50, match_types="아이템 개인전", start_date=start, end_date=end)

    rank_layout = [[], [], [], [], [], [], [], []]
    track = []
    rank = deepcopy(rank_layout)
    result = []
    s = "메이플 헤네시스 공원"

    print(match_list)


    for gametype, match in match_list.items():
        for game in match:
            n = game.playercount
            #if n not in [2, 4, 8]:
            #    continue
            i = 0
            #print('\t', rank)
            #print('\t', track)
            for player in game.detail.players:
                if len(rank[i]) == 0:
                    rank[i].insert(1, player.charactername)
                rank[i].insert(1, player.matchrank)
                i += 1
            track.insert(0, game.detail.track)
            if game.detail.track == s:
                if s[0] == 'W':
                    s = "월드 이탈리아 여행"
                elif s[0] == '월':
                    s = "메이플 헤네시스 공원"
                if len(track) != 0:
                    for i in rank:
                        i.append(point(i[1:]))
                    result.append([rank, track])
                    print(rank)
                    print(track)
                    track = []
                    rank = deepcopy(rank_layout)

    cnt = 1

    wb = Workbook()
    del wb['Sheet']
    for r, t in result:
        ws = wb.create_sheet(f'{start.split()[0].replace("-", ".")} INDV #{cnt}')
        cnt += 1
        ws.column_dimensions['A'].width = 30
        for i in range(1, len(t) + 1):
            ws["a" + str(i + 1)] = t[i-1]

        for i in range(8):
            for j in range(len(r[i])):
                if r[i][j] == -1:
                    ws["bcdefghi"[i] + str(j + 1)].font = f
                ws["bcdefghi"[i] + str(j + 1)] = str(r[i][j]) if r[i][j] != -1 else 'R'



    wb.save(f'./result/{start.split()[0].replace("-", ".")} INDV.xlsx')


def item_indv_search(nick, start, end):
    start = SetTime(start)
    end = SetTime(end)
    match_list = api.user(nick).getMatches(limit=50, match_types="아이템 개인전", start_date=start, end_date=end)

    rank_layout = [[], [], [], [], [], [], [], []]
    track = []
    rank = deepcopy(rank_layout)
    result = []

    print(match_list)


    for gametype, match in match_list.items():
        for game in match:
            n = game.playercount
            i = 0
            print('\t', rank)
            print('\t', track)
            for player in game.detail.players:
                if len(rank[i]) == 0:
                    rank[i].insert(1, player.charactername)
                rank[i].insert(1, player.matchrank)
                i += 1
            track.insert(0, game.detail.track)
        if len(track) != 0:
            for i in rank:
                i.append(item_point(i[1:]))
            result.append([rank, track])
            print(rank)
            print(track)
            track = []
            rank = deepcopy(rank_layout)

    cnt = 1

    wb = Workbook()
    del wb['Sheet']
    for r, t in result:
        ws = wb.create_sheet(f'{start.split()[0].replace("-", ".")} INDV #{cnt}')
        cnt += 1
        ws.column_dimensions['A'].width = 30
        for i in range(1, len(t) + 1):
            ws["a" + str(i + 1)] = t[i-1]

        for i in range(8):
            for j in range(len(r[i])):
                if r[i][j] == -1:
                    ws["bcdefghi"[i] + str(j + 1)].font = f
                ws["bcdefghi"[i] + str(j + 1)] = str(r[i][j]) if r[i][j] != -1 else 'R'



    wb.save(f'./result/{start.split()[0].replace("-", ".")} INDV2.xlsx')


#item_indv_search('', "2022-11-02 13:16:00", "2022-11-02 14:00:00")
#team_search('뻔나노', "2022-10-27 22:00:00", "2022-10-27 24:00:00", "MJJ", "MIX")
#team_search('뻔나노', "2022-11-15 22:30:00", "2022-11-16 00:30:00", "MJJ", "MIX")

#team_search('나우니마이트', "2022-11-20 22:00:00", "2022-11-20 24:00:00", "JB", "MJJ")
#team_search('나우니마이트', "2022-11-21 20:00:00", "2022-11-21 22:40:00", "JB", "S8ST")
#team_search('나우니마이트', "2022-11-25 23:00:00", "2022-11-26 00:30:00", "WHG", "MJJ")
#team_search('나우니마이트', "2022-11-27 21:00:00", "2022-11-27 22:45:00", "MIX", "MJJ")
#team_search('나우니마이트', "2022-11-27 22:46:00", "2022-11-27 24:00:00", "S8DR", "MIX")
#team_search('나우니마이트', "2022-11-29 23:00:00", "2022-11-30 00:10:00", "S8ST", "MJJ")
#team_search('나우니마이트', "2022-12-01 23:00:00", "2022-12-02 00:30:00", "S8ST", "MJJ")
#team_search('나우니마이트', "2022-12-02 20:00:00", "2022-12-02 24:00:00", "LPR", "NR")
#team_search('나우니마이트', "2022-12-04 21:05:00", "2022-12-04 22:30:00", "MJJ", "S8DR")
#team_search('나우니마이트', "2022-12-06 22:40:00", "2022-12-06 24:00:00", "MJJ", "S8DR")
#team_search('나우니마이트', "2022-12-07 20:15:00", "2022-12-07 22:10:00", "JB", "MJJ")
#team_search('나우니마이트', "2022-12-11 21:00:00", "2022-12-11 22:30:00", "MJJ", "WHG")
#team_search('나우니마이트', "2022-12-12 21:08:00", "2022-12-12 21:38:00", "S8ST", "MJJ2")
#team_search('l나노ll', "2022-12-12 21:38:00", "2022-12-12 22:40:00", "S8ST", "MJJ", dir="./result")